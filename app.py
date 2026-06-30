"""
AURIC Navigation Bar — REST API
================================

Reads Navigation_Bar_Final.xlsx and exposes its data as JSON via HTTP endpoints.

Run:
  pip install flask flask-cors openpyxl
  python app.py
  # Open http://localhost:5000

Endpoints:
  GET  /                              -- API documentation (HTML)
  GET  /api/health                    -- health check
  GET  /api/plot                      -- master plot info
  GET  /api/land                      -- all land module data
  GET  /api/land/entity               -- entity & plot info
  GET  /api/land/dates                -- key dates
  GET  /api/land/schedule             -- project development schedule
  GET  /api/land/compliance           -- compliance rules
  GET  /api/land/title-management     -- activity change, sub-let, sub-lease, etc.
  GET  /api/planning                  -- all planning module data
  GET  /api/planning/area-statement   -- land area statement
  GET  /api/planning/allotment        -- plot allotment status (was the pie chart)
  GET  /api/planning/plots-cc         -- plots with CC status
  GET  /api/planning/plots-oc         -- plots with OC status
  GET  /api/planning/bpas             -- BPAS commencement & occupancy
  GET  /api/planning/production       -- plots under production
  GET  /api/planning/employment       -- employment generation
  GET  /api/planning/fees             -- BPAS fees
  GET  /api/utility                   -- utility plot details + charges
  GET  /api/utility/charges           -- utility charge breakdown
  GET  /api/finance                   -- all finance subsections
  GET  /api/finance/land-payments
  GET  /api/finance/bpas-fees
  GET  /api/finance/technical
  GET  /api/finance/fsi
  GET  /api/finance/tree-cutting
  GET  /api/finance/temp-construction
  GET  /api/csr                       -- CSR activities
  GET  /api/all                       -- full dump of everything
"""

from flask import Flask, jsonify, request, render_template_string
from flask_cors import CORS
from openpyxl import load_workbook
from datetime import datetime, date
import os

app = Flask(__name__)
CORS(app)

EXCEL_PATH = os.environ.get("AURIC_XLSX", "Navigation_Bar_Final.xlsx")

# ============================================================
# Parser — reads workbook into a structured dict
# ============================================================

def _v(cell_value):
    """Normalize cell value for JSON output."""
    if isinstance(cell_value, (datetime, date)):
        return cell_value.strftime("%Y-%m-%d")
    if cell_value is None:
        return None
    if isinstance(cell_value, str) and cell_value.strip() == "—":
        return None
    return cell_value

def _read_kv_block(ws, start_row, end_row, key_col=2, val_col=3):
    """Read a key-value block where keys are in key_col and values in val_col."""
    out = {}
    for r in range(start_row, end_row + 1):
        k = ws.cell(row=r, column=key_col).value
        v = ws.cell(row=r, column=val_col).value
        if k and isinstance(k, str) and k.strip():
            out[k.strip()] = _v(v)
    return out

def _read_grid(ws, header_row, first_data_row, last_data_row, col_starts):
    """Read a grid table; returns list of dicts keyed by column headers."""
    headers = []
    for cs in col_starts:
        h = ws.cell(row=header_row, column=cs).value
        headers.append(h if h else f"col_{cs}")
    rows = []
    for r in range(first_data_row, last_data_row + 1):
        # Skip rows that are entirely empty
        row_vals = [_v(ws.cell(row=r, column=cs).value) for cs in col_starts]
        if all(x is None for x in row_vals):
            continue
        rows.append(dict(zip(headers, row_vals)))
    return rows

def _find_row(ws, text, col=2, start=1, end=None):
    """Find first row where column `col` contains the given text."""
    end = end or ws.max_row
    for r in range(start, end + 1):
        v = ws.cell(row=r, column=col).value
        if v and isinstance(v, str) and text.lower() in v.lower():
            return r
    return None

def parse_workbook(path):
    """Parse the entire workbook into a structured dictionary."""
    wb = load_workbook(path, data_only=True)
    ws = wb["AURIC Dashboard"]

    data = {
        "meta": {
            "source": os.path.basename(path),
            "parsed_at": datetime.utcnow().isoformat() + "Z",
            "location": "Bidkin · Sector 4 · Plot 33",
        }
    }

    # ---- Master Plot Info (rows ~6-22) ----
    master_start = _find_row(ws, "Master Plot Information") + 1
    # find the next section banner
    section_end = _find_row(ws, "1. LAND", start=master_start) - 1
    data["plot"] = _read_kv_block(ws, master_start, section_end)

    # ---- 1. LAND ----
    land = {}
    r = _find_row(ws, "1.1 Entity & Plot Info")
    end = _find_row(ws, "1.2 Key Dates") - 1
    land["entity"] = _read_kv_block(ws, r+1, end)

    r = _find_row(ws, "1.2 Key Dates")
    end = _find_row(ws, "1.3 Project Development") - 1
    land["dates"] = _read_kv_block(ws, r+1, end)

    r = _find_row(ws, "1.3 Project Development")
    end = _find_row(ws, "1.4 Compliance") - 1
    land["schedule"] = _read_grid(ws, r+1, r+2, end, [2, 3, 10, 12])

    r = _find_row(ws, "1.4 Compliance")
    end = _find_row(ws, "1.5 Land Use") - 1
    land["compliance"] = _read_kv_block(ws, r+1, end)

    # Title management subsections
    title_mgmt = {}
    for sub in ["1.5.1 Activity Change", "1.5.2 Final Amalgamation",
                "1.5.3 Sub-let", "1.5.4 Sub-lease", "1.5.5 Expansion",
                "1.5.6 Consent to Mortgage"]:
        sr = _find_row(ws, sub)
        if not sr:
            continue
        # Find next subsection or section
        next_r = None
        for nxt in ["1.5.2","1.5.3","1.5.4","1.5.5","1.5.6","2. PLANNING"]:
            if nxt > sub[:5]:
                cand = _find_row(ws, nxt, start=sr+1)
                if cand:
                    next_r = cand
                    break
        end_r = (next_r - 1) if next_r else sr + 10
        key = sub.split(" ", 1)[1].lower().replace(" ", "_").replace("-", "_")
        if "amalgamation" in sub:
            title_mgmt[key] = _read_grid(ws, sr+1, sr+2, end_r,
                                         [2, 6, 8, 10, 12])
        else:
            title_mgmt[key] = _read_kv_block(ws, sr+1, end_r)
    land["title_management"] = title_mgmt
    data["land"] = land

    # ---- 2. PLANNING ----
    planning = {}
    r = _find_row(ws, "2.1 Summary")
    end = _find_row(ws, "2.2 Land Area Statement") - 1
    planning["summary"] = _read_kv_block(ws, r+1, end)

    r = _find_row(ws, "2.2 Land Area Statement")
    end = _find_row(ws, "2.3 Building Project") - 1
    planning["area_statement"] = _read_grid(ws, r+1, r+2, end, [2, 3, 8, 11])

    # 2.4 allotment table (replaces former pie chart)
    r = _find_row(ws, "2.4 Plot Allotment Status")
    allotment_header = r + 1
    allotment_rows = _read_grid(ws, allotment_header, allotment_header+1,
                                 allotment_header+3, [2, 5, 7, 9])
    planning["allotment_status"] = allotment_rows

    # Plots-CC table
    r = _find_row(ws, "2.3 Building Project")
    end = _find_row(ws, "2.4 Plot Allotment Status") - 1
    planning["plots_cc"] = _read_grid(ws, r+1, r+3, end, [2, 4, 6, 8, 10, 12])

    # Plots-OC table
    r = _find_row(ws, "2.5 Plots Allotted (OC)")
    end = _find_row(ws, "2.6 BPAS") - 1
    planning["plots_oc"] = _read_grid(ws, r+1, r+3, end,
                                      [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13])

    # BPAS CC & OC
    r = _find_row(ws, "2.6 BPAS")
    end = _find_row(ws, "2.7 BPAS") - 1
    planning["bpas_cc"] = _read_grid(ws, r+2, r+3, end, [2, 6, 8, 10, 12])

    r = _find_row(ws, "2.7 BPAS")
    end = _find_row(ws, "2.8 Plots under Production") - 1
    planning["bpas_oc"] = _read_grid(ws, r+1, r+2, end, [2, 6, 8, 10, 12])

    # Production table
    r = _find_row(ws, "2.8 Plots under Production")
    end = _find_row(ws, "2.9 Employment") - 1
    planning["production"] = []
    for rr in range(r+3, end + 1):
        sector = _v(ws.cell(row=rr, column=2).value)
        if sector is None: continue
        planning["production"].append({
            "Sector": sector,
            "Industrial_Count": _v(ws.cell(row=rr, column=3).value),
            "Industrial_Area_Ha": _v(ws.cell(row=rr, column=5).value),
            "OpenSpace_Count": _v(ws.cell(row=rr, column=7).value),
            "OpenSpace_Area_Ha": _v(ws.cell(row=rr, column=9).value),
            "PT_Drains_Count": _v(ws.cell(row=rr, column=11).value),
            "PT_Drains_Area_Ha": _v(ws.cell(row=rr, column=13).value),
        })

    # Employment
    r = _find_row(ws, "2.9 Employment")
    end = _find_row(ws, "2.10 Fees") - 1
    planning["employment"] = []
    for rr in range(r+3, end + 1):
        sector = _v(ws.cell(row=rr, column=2).value)
        if sector is None: continue
        planning["employment"].append({
            "Sector": sector,
            "MSME": _v(ws.cell(row=rr, column=3).value),
            "Mega": _v(ws.cell(row=rr, column=7).value),
            "UltraMega": _v(ws.cell(row=rr, column=11).value),
        })

    # Fees grid
    r = _find_row(ws, "2.10 Fees")
    end = _find_row(ws, "3. UTILITY") - 1
    planning["bpas_fees_grid"] = _read_grid(ws, r+1, r+2, end,
                                            [2, 5, 6, 7, 8, 9, 10, 11, 12, 13])
    data["planning"] = planning

    # ---- 3. UTILITY ----
    utility = {}
    r = _find_row(ws, "3.1 Plot & Project Details")
    end = _find_row(ws, "3.2 Utility Charges") - 1
    utility["details"] = _read_kv_block(ws, r+1, end)

    r = _find_row(ws, "3.2 Utility Charges")
    end = _find_row(ws, "4. FINANCE") - 1
    charges = {}
    for rr in range(r+2, end + 1):
        k = ws.cell(row=rr, column=2).value
        v = ws.cell(row=rr, column=8).value
        if k and isinstance(k, str):
            charges[k.strip()] = _v(v)
    utility["charges"] = charges
    data["utility"] = utility

    # ---- 4. FINANCE ----
    finance = {}
    for sub_id, sub_name, key in [
        ("4.1 Land Payments", "Land Payments", "land_payments"),
        ("4.3 BPAS Fees", "BPAS Fees", "bpas_fees"),
        ("4.4 FSI Payment", "FSI Payment", "fsi"),
        ("4.5 Tree Cutting Fee", "Tree Cutting Fee", "tree_cutting"),
        ("4.6 Temporary Construction Fees", "Temporary Construction Fees", "temp_construction"),
    ]:
        sr = _find_row(ws, sub_id)
        if not sr: continue
        # Find next subsection
        next_subs = ["4.2 Technical","4.3 BPAS","4.4 FSI","4.5 Tree","4.6 Temporary","5. CSR"]
        next_r = None
        for ns in next_subs:
            if sub_id[:3] < ns[:3] or (sub_id[:3] == ns[:3] and sub_id < ns):
                cand = _find_row(ws, ns, start=sr+1)
                if cand:
                    next_r = cand; break
        end_r = (next_r - 1) if next_r else sr + 15
        items = {}
        for rr in range(sr+2, end_r + 1):
            k = ws.cell(row=rr, column=2).value
            v = ws.cell(row=rr, column=10).value
            if k and isinstance(k, str) and k.strip() and "Total" not in k:
                items[k.strip()] = _v(v)
        finance[key] = items

    # Technical
    r = _find_row(ws, "4.2 Technical")
    if r:
        end = _find_row(ws, "4.3 BPAS") - 1
        finance["technical"] = _read_kv_block(ws, r+1, end)
    data["finance"] = finance

    # ---- 5. CSR ----
    r = _find_row(ws, "5.1 CSR")
    if r:
        end = _find_row(ws, "6. OTHERS") - 1 if _find_row(ws, "6. OTHERS") else ws.max_row
        data["csr"] = _read_grid(ws, r+1, r+2, end, [2, 5, 9, 11, 12])
    else:
        data["csr"] = []

    return data

# Parse once at startup
print(f"Loading {EXCEL_PATH}...")
DATA = parse_workbook(EXCEL_PATH)
print(f"Loaded {sum(1 for _ in DATA)} top-level keys")

# ============================================================
# Routes
# ============================================================

DOCS_HTML = """
<!DOCTYPE html>
<html><head><title>AURIC Navigation Bar API</title>
<style>
  body { font-family: -apple-system, Segoe UI, sans-serif; max-width: 900px; margin: 40px auto; padding: 0 20px; color: #1e293b; }
  h1 { color: #0F4C81; border-bottom: 3px solid #0F4C81; padding-bottom: 8px; }
  h2 { color: #1E5A8E; margin-top: 32px; }
  .endpoint { background: #f1f5f9; padding: 12px 16px; border-radius: 6px; margin: 6px 0; font-family: Consolas, monospace; font-size: 14px; }
  .endpoint a { color: #0F4C81; text-decoration: none; font-weight: bold; }
  .endpoint a:hover { text-decoration: underline; }
  .desc { color: #64748b; font-size: 13px; margin-left: 8px; }
  .badge { display: inline-block; background: #0F4C81; color: white; padding: 2px 8px; border-radius: 3px; font-size: 11px; font-weight: bold; margin-right: 8px; }
</style></head>
<body>
<h1>🏭 AURIC Navigation Bar — REST API</h1>
<p><strong>Source:</strong> Navigation_Bar_Final.xlsx · <strong>Location:</strong> Bidkin · Sector 4 · Plot 33</p>
<p>This API serves all data from the AURIC dashboard workbook as JSON. All endpoints support CORS and return <code>application/json</code>.</p>

<h2>System</h2>
<div class="endpoint"><span class="badge">GET</span><a href="/api/health">/api/health</a><span class="desc">Health check</span></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/all">/api/all</a><span class="desc">Complete data dump</span></div>

<h2>Master Plot</h2>
<div class="endpoint"><span class="badge">GET</span><a href="/api/plot">/api/plot</a><span class="desc">Master plot information</span></div>

<h2>1. Land</h2>
<div class="endpoint"><span class="badge">GET</span><a href="/api/land">/api/land</a><span class="desc">All land data</span></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/land/entity">/api/land/entity</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/land/dates">/api/land/dates</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/land/schedule">/api/land/schedule</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/land/compliance">/api/land/compliance</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/land/title-management">/api/land/title-management</a></div>

<h2>2. Planning</h2>
<div class="endpoint"><span class="badge">GET</span><a href="/api/planning">/api/planning</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/planning/area-statement">/api/planning/area-statement</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/planning/allotment">/api/planning/allotment</a><span class="desc">Former pie chart — now a table</span></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/planning/plots-cc">/api/planning/plots-cc</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/planning/plots-oc">/api/planning/plots-oc</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/planning/bpas">/api/planning/bpas</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/planning/production">/api/planning/production</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/planning/employment">/api/planning/employment</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/planning/fees">/api/planning/fees</a></div>

<h2>3. Utility</h2>
<div class="endpoint"><span class="badge">GET</span><a href="/api/utility">/api/utility</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/utility/charges">/api/utility/charges</a></div>

<h2>4. Finance</h2>
<div class="endpoint"><span class="badge">GET</span><a href="/api/finance">/api/finance</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/finance/land-payments">/api/finance/land-payments</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/finance/bpas-fees">/api/finance/bpas-fees</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/finance/technical">/api/finance/technical</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/finance/fsi">/api/finance/fsi</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/finance/tree-cutting">/api/finance/tree-cutting</a></div>
<div class="endpoint"><span class="badge">GET</span><a href="/api/finance/temp-construction">/api/finance/temp-construction</a></div>

<h2>5. CSR</h2>
<div class="endpoint"><span class="badge">GET</span><a href="/api/csr">/api/csr</a></div>

<hr style="margin-top:40px;border:none;border-top:1px solid #e2e8f0;">
<p style="color:#94a3b8;font-size:12px;text-align:center;">AURIC Navigation Bar REST API · Built with Flask</p>
</body></html>
"""

@app.route("/")
def root():
    return DOCS_HTML

@app.route("/api/health")
def health():
    return jsonify(status="ok", parsed_at=DATA["meta"]["parsed_at"], source=DATA["meta"]["source"])

@app.route("/api/all")
def all_data():
    return jsonify(DATA)

@app.route("/api/plot")
def plot():
    return jsonify(DATA["plot"])

# Land
@app.route("/api/land")
def land(): return jsonify(DATA["land"])
@app.route("/api/land/entity")
def land_entity(): return jsonify(DATA["land"]["entity"])
@app.route("/api/land/dates")
def land_dates(): return jsonify(DATA["land"]["dates"])
@app.route("/api/land/schedule")
def land_schedule(): return jsonify(DATA["land"]["schedule"])
@app.route("/api/land/compliance")
def land_compliance(): return jsonify(DATA["land"]["compliance"])
@app.route("/api/land/title-management")
def land_title(): return jsonify(DATA["land"]["title_management"])

# Planning
@app.route("/api/planning")
def planning(): return jsonify(DATA["planning"])
@app.route("/api/planning/area-statement")
def planning_area(): return jsonify(DATA["planning"]["area_statement"])
@app.route("/api/planning/allotment")
def planning_allotment(): return jsonify(DATA["planning"]["allotment_status"])
@app.route("/api/planning/plots-cc")
def planning_cc(): return jsonify(DATA["planning"]["plots_cc"])
@app.route("/api/planning/plots-oc")
def planning_oc(): return jsonify(DATA["planning"]["plots_oc"])
@app.route("/api/planning/bpas")
def planning_bpas(): return jsonify(commencement=DATA["planning"]["bpas_cc"],
                                     occupancy=DATA["planning"]["bpas_oc"])
@app.route("/api/planning/production")
def planning_prod(): return jsonify(DATA["planning"]["production"])
@app.route("/api/planning/employment")
def planning_emp(): return jsonify(DATA["planning"]["employment"])
@app.route("/api/planning/fees")
def planning_fees(): return jsonify(DATA["planning"]["bpas_fees_grid"])

# Utility
@app.route("/api/utility")
def utility(): return jsonify(DATA["utility"])
@app.route("/api/utility/charges")
def utility_charges(): return jsonify(DATA["utility"]["charges"])

# Finance
@app.route("/api/finance")
def finance(): return jsonify(DATA["finance"])
@app.route("/api/finance/land-payments")
def fin_land(): return jsonify(DATA["finance"].get("land_payments", {}))
@app.route("/api/finance/bpas-fees")
def fin_bpas(): return jsonify(DATA["finance"].get("bpas_fees", {}))
@app.route("/api/finance/technical")
def fin_tech(): return jsonify(DATA["finance"].get("technical", {}))
@app.route("/api/finance/fsi")
def fin_fsi(): return jsonify(DATA["finance"].get("fsi", {}))
@app.route("/api/finance/tree-cutting")
def fin_tree(): return jsonify(DATA["finance"].get("tree_cutting", {}))
@app.route("/api/finance/temp-construction")
def fin_temp(): return jsonify(DATA["finance"].get("temp_construction", {}))

# CSR
@app.route("/api/csr")
def csr(): return jsonify(DATA["csr"])

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"\nAPI running at: http://0.0.0.0:{port}")
    print("Open in browser for documentation.\n")
    app.run(host="0.0.0.0", port=port, debug=False)
